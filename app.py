import urllib2
from bs4 import BeautifulSoup
from openpyxl import load_workbook


productlists = []
productnames = ['PRODUCT NAME']
special_pricelists = ['Special PRICE']
old_pricelists = ['Old PRICE']
minimal_pricelists = ['MINIMAL PRICE']

scraping_url = "https://www.frankana.de/de/sale/monatsangebot.html"
page = urllib2.urlopen(scraping_url)
soup = BeautifulSoup(page, 'html.parser')
i = 0
wb = load_workbook('/Volumes/Work/study/python/WebScraping/result.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']

# ---------- product name list scraping ---------------- #

productlists = soup.find_all('h2', class_='product-name')
for _product in productlists:
    i = i + 1
    product = _product.find("a").get_text()

    productnames.insert(i, product)


# ---------- product price scraping -------------------- #
special_price_p = soup.select("div.price-box p.special-price span.price")
old_price_p = soup.select("div.price-box p.old-price span.price")
minimal_price_p = soup.select("div.price-box p.minimal-price span.price")


for j, tag in enumerate(special_price_p):

    special_pricelists.append(special_price_p[j].text)
    minimal_pricelists.append(" --- ")

for k, tag in enumerate(old_price_p):

    old_pricelists.append(old_price_p[k].text)

for t, tag in enumerate(minimal_price_p):
    minimal_pricelists.append(minimal_price_p[t].text)

# -----------  save scraping data in excel sheet --------#
for i, tag in enumerate(productnames):
    sheet.cell(row=i+1, column=1).value = productnames[i]

for i, tag in enumerate(special_pricelists):

    sheet.cell(row=i+1, column=2).value = special_pricelists[i]

for i, tag in enumerate(old_pricelists):
    sheet.cell(row=i+1, column=3).value = old_pricelists[i]

for i, tag in enumerate(minimal_pricelists):

    sheet.cell(row=i+1, column=4).value = minimal_pricelists[i]

wb.save('/Volumes/Work/study/python/WebScraping/result.xlsx')


print "productnames"
print productnames

print "special_pricelists"
print special_pricelists

print "old_pricelists"
print old_pricelists





